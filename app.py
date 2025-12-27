import pandas as pd
import sqlite3
import uuid
import os
import glob
import io
import numpy as np
from flask import Flask, render_template, request, redirect, url_for, session, g, send_from_directory, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Alignment
"""
==================================================================================
NOTA TÉCNICA E METODOLÓGICA (DNIT 008/2003 - PRO)
==================================================================================

1. APROXIMAÇÃO TÉCNICA AUTOMATIZADA (CÁLCULO POR ÁREA):
   Este software implementa uma adaptação computacional do método de "Levantamento
   Visual Contínuo". Enquanto a norma original baseia-se na avaliação visual da
   % de extensão do segmento afetada, este algoritmo utiliza os dados quantitativos
   de área (m²) presentes no inventário para calcular uma "Extensão Equivalente".
   
   Metodologia Adotada:
   - Trincas e Deformações: O cálculo da frequência (%D) é realizado através da
     razão entre a [Soma das Áreas Afetadas] e a [Área Total do Segmento].
     Esta abordagem converte o dado de área em um indicador percentual compatível
     com as tabelas de classificação da norma.
   - Panelas e Remendos: Mantém-se a contagem absoluta de ocorrências (unidades/km),
     estritamente conforme a Tabela 1 da norma.

2. MAPEAMENTO DE COLUNAS E CONJUNTOS NORMATIVOS:
   O dicionário de configuração de colunas ('INDICES') não representa campos
   aleatórios, mas sim os agrupamentos de defeitos definidos no Anexo A da norma.
   Independentemente da posição na planilha, os dados são agregados nos seguintes
   conjuntos lógicos para cálculo do IGGE:
   - Grupo Trincas (Pt): Fissuras, Trincas Isoladas, Jacaré (J/JE), Bloco (TB/TBE).
   - Grupo Deformações (Poap): Afundamentos (ALP/ATP/ALC/ATC) e Ondulações.
   - Grupo Panelas (Ppr): Soma de Panelas (P) e Remendos (R) Superficiais/Profundos.

==================================================================================
"""

app = Flask(__name__)
app.config['SECRET_KEY'] = 'cf83e1357eefb8bdf1542850d66d8007d620e4050b5715dc83f4a921d36ce9ce'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

DATABASE = 'projeto.db'

LARGURA_FAIXA = 7       # 2 Faixas (LD + LE)
COMPRIMENTO_ESTACA = 20.0 
AREA_ESTACA = LARGURA_FAIXA * COMPRIMENTO_ESTACA # 140.0 m² (A_i)

INDICES = {
    'km': 0,
    'trincas': [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34],
    'deformacoes': [12, 13, 14, 15, 16, 35, 36, 37, 38, 39],
    'panelas_remendos': [17, 40, 21, 44]
}

PESOS = {'Trincas': 0.35, 'Deformacoes': 0.60, 'Panelas': 0.70}

FATORES_GRAV = {
    'Trincas': {'A': 0.65, 'M': 0.45, 'B': 0.30},
    'Deformacoes': {'A': 1.00, 'M': 0.70, 'B': 0.60},
    'Panelas': {'A': 1.00, 'M': 0.80, 'B': 0.70}
}

def limpar_uploads_ao_iniciar():
    pasta = app.config['UPLOAD_FOLDER']
    if not os.path.exists(pasta):
        os.makedirs(pasta)
        return
    arquivos = glob.glob(os.path.join(pasta, '*'))
    for arquivo in arquivos:
        try: os.remove(arquivo)
        except: pass

limpar_uploads_ao_iniciar()

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None: db.close()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def normalizar_para_float(valor):
    if pd.isna(valor) or valor == '': return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    valor_str = str(valor).strip().replace(',', '.')
    try: return float(valor_str)
    except ValueError: return 0.0

def calcular_igge_pro008(df, upload_id, metodo_panelas):
    db = get_db()
    cursor = db.cursor()

    # 1. PROCESSAMENTO INICIAL
    df_proc = pd.DataFrame()
    max_col = df.shape[1]
    
    idx_km = INDICES['km']
    if idx_km < max_col:
        df_proc['km'] = df.iloc[:, idx_km].apply(normalizar_para_float)
        df_proc['km_segmento'] = df_proc['km'].apply(lambda x: int(x))
    else:
        raise ValueError("A planilha não tem a coluna 0 (KM).")

    def processar_grupo(row_idx, lista_indices, modo='soma_real'):
        soma = 0.0
        for col_idx in lista_indices:
            if col_idx < max_col:
                val = normalizar_para_float(df.iloc[row_idx, col_idx])
                if modo == 'soma_real':
                    soma += val
                elif modo == 'binario':
                    soma += 1 if val > 0 else 0
        return soma

    areas_trinca = []
    areas_deform = []
    qtds_panela = []

    modo_panela_escolhido = 'soma_real' if metodo_panelas == 'contagem' else 'binario'

    for i in range(len(df)):
        areas_trinca.append(processar_grupo(i, INDICES['trincas'], modo='soma_real'))
        
        areas_deform.append(processar_grupo(i, INDICES['deformacoes'], modo='soma_real'))
        
        qtds_panela.append(processar_grupo(i, INDICES['panelas_remendos'], modo=modo_panela_escolhido))


    df_proc['area_trinca'] = areas_trinca
    df_proc['area_deform'] = areas_deform
    df_proc['qtd_panelas'] = qtds_panela

    # 2. AGRUPAMENTO POR SEGMENTO (LÓGICA HÍBRIDA)
    df_resumo = df_proc.groupby('km_segmento').apply(lambda x: pd.Series({
        'total_estacas': len(x),
        'area_total_segmento': len(x) * AREA_ESTACA, 
        
        'soma_area_trinca': x['area_trinca'].sum(),
        'soma_area_deform': x['area_deform'].sum(),
        
        'qtd_total_panelas_remendos': round(x['qtd_panelas'].sum(), 0)
    }), include_groups=False).reset_index()

    # 3. CÁLCULO DAS FREQUÊNCIAS (% e Quantidade)
    
    
    df_resumo['pct_trincas'] = (df_resumo['soma_area_trinca'] / df_resumo['area_total_segmento']) * 100
    df_resumo['freq_trincas'] = df_resumo['pct_trincas'].apply(
        lambda p: 'A' if p >= 50 else ('M' if p > 10 else 'B')
    )
    
    df_resumo['pct_deform'] = (df_resumo['soma_area_deform'] / df_resumo['area_total_segmento']) * 100
    df_resumo['freq_deform'] = df_resumo['pct_deform'].apply(
        lambda p: 'A' if p >= 50 else ('M' if p > 10 else 'B')
    )
    
    df_resumo['freq_panelas'] = df_resumo['qtd_total_panelas_remendos'].apply(
        lambda q: 'A' if q >= 5 else ('M' if q > 2 else 'B')
    )

    # 4. FATORES E IGGE
    df_resumo['pt'] = df_resumo['freq_trincas'].map(FATORES_GRAV['Trincas'])
    df_resumo['poap'] = df_resumo['freq_deform'].map(FATORES_GRAV['Deformacoes'])
    df_resumo['ppr'] = df_resumo['freq_panelas'].map(FATORES_GRAV['Panelas'])

    df_resumo['igge'] = (
            (df_resumo['pt'] * df_resumo['pct_trincas']) + 
            (df_resumo['poap'] * df_resumo['pct_deform']) + 
            (df_resumo['ppr'] * df_resumo['qtd_total_panelas_remendos'])
        )
    
    def estimar_icpf(igge):
        if igge <= 5: return 5
        elif igge <= 15: return 4
        elif igge <= 40: return 3
        elif igge <= 70: return 2
        else: return 1

    df_resumo['icpf'] = df_resumo['igge'].apply(estimar_icpf)

    # 5. CONCEITO E IES
    def determinar_ies_conceito(row):
        igge = row['igge']
        icpf = row['icpf']

        if igge <= 20:
            if icpf > 3.5: return 0, 'Ótimo'   # Código A
            else:          return 1, 'Bom'     # Código B
            
        elif igge <= 40: # 20 < IGGE <= 40
            if icpf > 3.5: return 2, 'Bom'     # Código B 
            else:          return 3, 'Regular' # Código C
            
        elif igge <= 60: # 40 < IGGE <= 60
            if icpf > 2.5: return 4, 'Regular' # Código C
            else:          return 5, 'Ruim'    # Código D
            
        elif igge <= 90: # 60 < IGGE <= 90
            if icpf > 2.5: return 7, 'Ruim'    # Código D
            else:          return 8, 'Péssimo' # Código E
            
        else: # IGGE > 90
            return 10, 'Péssimo'               # Código E
        
    df_resumo[['ies', 'conceito']] = df_resumo.apply(determinar_ies_conceito, axis=1, result_type='expand')

    # 6. SALVAR NO BANCO
    cursor.execute("DELETE FROM resultados_pro008 WHERE upload_id = ?", (upload_id,))
    
    dados_insert = []
    for _, row in df_resumo.iterrows():
        dados_insert.append((
            upload_id, row['km_segmento'], row['km_segmento']+1,
            int(row['total_estacas']),
            
            round(row['soma_area_trinca'], 2), row['pct_trincas'],
            round(row['soma_area_deform'], 2), row['pct_deform'],
            
            int(row['qtd_total_panelas_remendos']),
            
            row['freq_trincas'], row['freq_deform'], row['freq_panelas'],
            row['pt'], row['poap'], row['ppr'],
            row['igge'], row['icpf'], row['ies'], row['conceito']
        ))

    sql = """INSERT INTO resultados_pro008 
                (upload_id, km_inicial, km_final, total_estacas, 
                qtd_trincas, pct_trincas, qtd_deformacoes, pct_deformacoes, qtd_panelas,
                freq_trincas, freq_deformacoes, freq_panelas,
                grav_trincas, grav_deformacoes, grav_panelas,
                igge, icpf, ies, conceito)  -- Adicionado icpf aqui
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
    cursor.executemany(sql, dados_insert)
    db.commit()

# --- ROTAS FLASK ---
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files: return redirect(request.url)
        file = request.files['file']
        try: linha_inicial = int(request.form.get('linha_inicial', 1))
        except: linha_inicial = 1

        metodo_panelas = request.form.get('metodo_panelas', 'incidencia')

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            uid = str(uuid.uuid4())
            session['upload_id'] = uid
            try:
                df = pd.read_excel(filepath, header=None, skiprows=linha_inicial-1, dtype=object)
                calcular_igge_pro008(df, uid, metodo_panelas)
                return redirect(url_for('relatorio', id=uid))
            except Exception as e:
                return render_template('index.html', error=str(e))
    return render_template('index.html')

@app.route('/relatorio/<id>')
def relatorio(id):
    db = get_db()
    res = db.execute("SELECT * FROM resultados_pro008 WHERE upload_id = ? ORDER BY km_inicial", (id,)).fetchall()
    chart_data = {
        'labels': [f"km {r['km_inicial']}" for r in res],
        'datasets': [
            {
                'label': 'IGGE (Gravidade Global)',
                'data': [r['igge'] for r in res],
                'backgroundColor': 'rgba(13, 110, 253, 0.5)',
                'borderColor': 'rgba(13, 110, 253, 1)',
                'borderWidth': 1,
                'type': 'bar', 
                'yAxisID': 'y', 
                'order': 2
            },
            {
                'label': 'IES (Estado da Superfície)',
                'data': [r['ies'] for r in res],
                'backgroundColor': 'rgba(255, 193, 7, 1)', 
                'borderColor': 'rgba(255, 193, 7, 1)',
                'borderWidth': 3,
                'type': 'line',
                'tension': 0.3, 
                'yAxisID': 'y1', 
                'order': 1
            },
            {
                'label': 'ICPF (Condição Estimada)',
                'data': [r['icpf'] for r in res],
                'backgroundColor': 'rgba(25, 135, 84, 1)', # Verde
                'borderColor': 'rgba(25, 135, 84, 1)',
                'borderWidth': 3,
                'type': 'line',
                'tension': 0.3,
                'yAxisID': 'y1',
                'order': 0
            }
        ]
    }
    return render_template('relatorio.html', resultados=res, chart_data=chart_data)

@app.route('/download_modelo')
def download_modelo():
    try:
        return send_from_directory(directory='static', path='modelo_padrao.xlsx', as_attachment=True)
    except FileNotFoundError:
        return "ERRO: Arquivo 'modelo_padrao.xlsx' não encontrado na pasta 'static'."
    
@app.route('/exportar_relatorio/<id>')
def exportar_relatorio(id):
    db = get_db()
    resultados = db.execute("SELECT * FROM resultados_pro008 WHERE upload_id = ? ORDER BY km_inicial", (id,)).fetchall()
    
    if not resultados:
        return "Sem dados para exportar.", 404

    caminho_template = os.path.join(app.config['UPLOAD_FOLDER'], '../static/template_exportacao.xlsx')
    
    wb = load_workbook(caminho_template)
    ws_c = wb['anexo_c'] 
    ws_d = wb['anexo_d']
    
    linha_atual = 10
    total_linhas = len(resultados)

    mapa_codigo_ies = {
            0: 'A', 
            1: 'B', 2: 'B', 
            3: 'C', 4: 'C', 
            5: 'D', 7: 'D', 
            8: 'E', 10: 'E'
        }
    
    for i, row in enumerate(resultados):
        km_ini = row['km_inicial']
        eh_ultimo = (i == total_linhas - 1)
        
        if eh_ultimo:
            extensao_real = row['total_estacas'] * 0.02
            km_fim = km_ini + extensao_real
        else:
            km_fim = row['km_inicial'] + 1.0

        estaca_inicial = round(km_ini / 0.02, 0)
        estaca_final = round(km_fim / 0.02, 0)
        extensao = km_fim - km_ini

        fat_trincas = row['grav_trincas'] * row['pct_trincas']
        fat_deform = row['grav_deformacoes'] * row['pct_deformacoes']
        fat_panelas = row['grav_panelas'] * row['qtd_panelas']

        # --- 3. PREENCHIMENTO ANEXO C ---
        ws_c.cell(row=linha_atual, column=1).value = km_ini
        ws_c.cell(row=linha_atual, column=2).value = estaca_inicial
        ws_c.cell(row=linha_atual, column=3).value = estaca_final
        ws_c.cell(row=linha_atual, column=4).value = km_ini
        ws_c.cell(row=linha_atual, column=5).value = km_fim
        ws_c.cell(row=linha_atual, column=6).value = extensao
        ws_c.cell(row=linha_atual, column=7).value = row['pct_trincas']
        ws_c.cell(row=linha_atual, column=8).value = row['grav_trincas']
        ws_c.cell(row=linha_atual, column=9).value = fat_trincas
        ws_c.cell(row=linha_atual, column=10).value = row['pct_deformacoes']
        ws_c.cell(row=linha_atual, column=11).value = row['grav_deformacoes']
        ws_c.cell(row=linha_atual, column=12).value = fat_deform
        ws_c.cell(row=linha_atual, column=13).value = row['qtd_panelas']
        ws_c.cell(row=linha_atual, column=14).value = row['grav_panelas']
        ws_c.cell(row=linha_atual, column=15).value = fat_panelas
        ws_c.cell(row=linha_atual, column=16).value = row['igge']

        for col in range(1, 17):
            cell = ws_c.cell(row=linha_atual, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if col in [4, 5, 6, 8, 9, 11, 12, 15]: 
                cell.number_format = '0.00'
    
        # --- 3. PREENCHIMENTO ANEXO D ---
        ws_d.cell(row=linha_atual, column=1).value = km_ini 
        ws_d.cell(row=linha_atual, column=2).value = estaca_inicial 
        ws_d.cell(row=linha_atual, column=3).value = estaca_final 
        ws_d.cell(row=linha_atual, column=4).value = km_ini 
        ws_d.cell(row=linha_atual, column=5).value = km_fim
        ws_d.cell(row=linha_atual, column=6).value = extensao 
        ws_d.cell(row=linha_atual, column=7).value = row['icpf']
        ws_d.cell(row=linha_atual, column=8).value = row['igge'] 
        ws_d.cell(row=linha_atual, column=9).value = row['ies']     
        codigo_ies = mapa_codigo_ies.get(row['ies'], '-')
        ws_d.cell(row=linha_atual, column=10).value = codigo_ies
        ws_d.cell(row=linha_atual, column=11).value = row['conceito']

        for col in range(1, 12):
            cell = ws_d.cell(row=linha_atual, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col in [4, 5, 6]: cell.number_format = '0.00' 
        linha_atual += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Relatorio_DNIT{id}.xlsx'
    )

if __name__ == '__main__':
    app.run(debug=True)